#!/usr/bin/env python3
"""
Apply Contact Fixes — Wendet Auto-Korrekturen auf Apple Contacts an
und merged Datalake-Kontakte ins Adressbuch.

Teil 1: AUTO-KORREKTUREN (Display-Name-Fixes aus der Mismatch-Analyse)
        → Aktualisiert ZFIRSTNAME/ZLASTNAME in AddressBook SQLite

Teil 2: MERGE DATALAKE → APPLE CONTACTS
        → Neue Kontakte aus contacts.json (die nicht schon per E-Mail
          existieren) werden ins Adressbuch eingefuegt.

BACKUP wurde separat erstellt unter:
  ~/AssistantDev/backups/addressbook_YYYYMMDD_HHMMSS/

Aufruf: python3 ~/AssistantDev/scripts/apply_contact_fixes.py
"""

import os
import sys
import re
import json
import glob
import uuid
import sqlite3
import shutil
import datetime

# ── Pfade ────────────────────────────────────────────────────────────────────

ADDRESSBOOK_DIR = os.path.expanduser("~/Library/Application Support/AddressBook")
DATALAKE = os.path.expanduser(
    "~/Library/Mobile Documents/com~apple~CloudDocs/Downloads shared/claude_datalake"
)
EXCEL_PATH = os.path.expanduser(
    "~/Library/Mobile Documents/com~apple~CloudDocs/Downloads shared/"
    "claude_outputs/contact_mismatch_2026-04-13.xlsx"
)

# AddressBook SQLite Konstanten (aus Schema-Analyse)
Z_ENT_RECORD = 22   # ZABCDRECORD entity type for a Person
Z_ENT_EMAIL = 11     # ZABCDEMAILADDRESS entity type
Z_ENT_PHONE = 15     # ZABCDPHONENUMBER entity type

AGENTS = ["signicat", "trustedcarrier", "privat", "standard"]

# ── Helpers ──────────────────────────────────────────────────────────────────

def split_name(display_name):
    """Splittet einen Display-Namen in (firstname, lastname).
    'Vorname Nachname' → ('Vorname', 'Nachname')
    'email@domain.com' → ('', 'email@domain.com')  # E-Mail bleibt im Nachnamen
    """
    if not display_name:
        return "", ""
    if "@" in display_name:
        # E-Mail als Display-Name → kein Vorname, E-Mail als "last name"
        return "", display_name
    parts = display_name.strip().split(None, 1)
    if len(parts) == 2:
        return parts[0], parts[1]
    if len(parts) == 1:
        return "", parts[0]
    return "", ""


def get_all_dbs():
    """Gibt alle AddressBook-Datenbanken zurueck die E-Mails enthalten."""
    dbs = glob.glob(os.path.join(ADDRESSBOOK_DIR, "Sources/*/AddressBook-v22.abcddb"))
    result = []
    for db in dbs:
        try:
            conn = sqlite3.connect(db)
            cur = conn.cursor()
            cur.execute("SELECT COUNT(*) FROM ZABCDEMAILADDRESS WHERE ZADDRESS IS NOT NULL")
            n = cur.fetchone()[0]
            conn.close()
            if n > 0:
                result.append(db)
        except Exception:
            pass
    return result


def build_email_index(db_path):
    """Baut einen Index email → (Z_PK, firstname, lastname, org) fuer eine DB."""
    conn = sqlite3.connect(db_path)
    cur = conn.cursor()
    cur.execute("""
        SELECT e.ZADDRESS, r.Z_PK, r.ZFIRSTNAME, r.ZLASTNAME, r.ZORGANIZATION
        FROM ZABCDEMAILADDRESS e
        JOIN ZABCDRECORD r ON e.ZOWNER = r.Z_PK
        WHERE e.ZADDRESS IS NOT NULL
    """)
    index = {}
    for addr, pk, fn, ln, org in cur.fetchall():
        if addr:
            index[addr.lower().strip()] = {
                "pk": pk, "firstname": fn or "", "lastname": ln or "",
                "org": org or "", "db": db_path,
            }
    conn.close()
    return index


# ═══════════════════════════════════════════════════════════════════════════
# Teil 1: AUTO-KORREKTUREN anwenden
# ═══════════════════════════════════════════════════════════════════════════

def apply_display_name_fixes():
    print("=" * 55)
    print("  TEIL 1: Display-Name-Korrekturen anwenden")
    print("=" * 55)

    try:
        from openpyxl import load_workbook
        wb = load_workbook(EXCEL_PATH)
    except Exception as e:
        print(f"FEHLER: Kann Excel nicht laden: {e}")
        return 0

    ws = wb["AUTO_KORREKTUREN"]
    corrections = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        source = row[0] or ""
        new_display = row[1] or ""
        email = (row[2] or "").lower().strip()
        org = row[4] or ""
        note = row[7] or ""
        if not email or not new_display:
            continue
        corrections.append({
            "email": email,
            "new_display": new_display,
            "org": org,
            "note": note,
        })

    print(f"  {len(corrections)} Korrekturen aus Excel geladen")

    # E-Mail-Index ueber alle DBs bauen
    dbs = get_all_dbs()
    print(f"  {len(dbs)} AddressBook-Datenbanken mit E-Mails")
    full_index = {}
    for db in dbs:
        idx = build_email_index(db)
        for email, info in idx.items():
            if email not in full_index:
                full_index[email] = info

    print(f"  {len(full_index)} E-Mail-Adressen im Index")

    applied = 0
    skipped = 0
    not_found = 0

    # Aenderungen pro DB sammeln
    changes_per_db = {}  # db_path -> list of (pk, firstname, lastname, org)

    for corr in corrections:
        email = corr["email"]
        info = full_index.get(email)
        if not info:
            not_found += 1
            continue

        new_fn, new_ln = split_name(corr["new_display"])
        old_fn = info["firstname"]
        old_ln = info["lastname"]

        # Nur aendern wenn tatsaechlich anders
        if new_fn == old_fn and new_ln == old_ln:
            skipped += 1
            continue

        db = info["db"]
        changes_per_db.setdefault(db, []).append({
            "pk": info["pk"],
            "new_fn": new_fn,
            "new_ln": new_ln,
            "new_org": corr["org"] if corr["org"] and not info["org"] else None,
            "email": email,
        })
        applied += 1

    # Aenderungen in die DBs schreiben
    for db_path, changes in changes_per_db.items():
        src_id = os.path.basename(os.path.dirname(db_path))[:12]
        print(f"\n  DB {src_id}: {len(changes)} Aenderungen")
        conn = sqlite3.connect(db_path)
        cur = conn.cursor()
        for ch in changes:
            now = datetime.datetime.now().timestamp() - 978307200  # CoreData epoch
            cur.execute("""
                UPDATE ZABCDRECORD
                SET ZFIRSTNAME = ?, ZLASTNAME = ?, ZMODIFICATIONDATE = ?
                WHERE Z_PK = ?
            """, (ch["new_fn"] or None, ch["new_ln"] or None, now, ch["pk"]))
            # Organisation nur setzen wenn bisher leer
            if ch["new_org"]:
                cur.execute("""
                    UPDATE ZABCDRECORD SET ZORGANIZATION = ?
                    WHERE Z_PK = ? AND (ZORGANIZATION IS NULL OR ZORGANIZATION = '')
                """, (ch["new_org"], ch["pk"]))
        conn.commit()
        conn.close()

    print(f"\n  Angewendet: {applied}")
    print(f"  Uebersprungen (schon korrekt): {skipped}")
    print(f"  Nicht im Adressbuch gefunden: {not_found}")
    return applied


# ═══════════════════════════════════════════════════════════════════════════
# Teil 2: DATALAKE-KONTAKTE MERGEN
# ═══════════════════════════════════════════════════════════════════════════

def merge_datalake_contacts():
    print("\n" + "=" * 55)
    print("  TEIL 2: Datalake-Kontakte ins Adressbuch mergen")
    print("=" * 55)

    # Alle Datalake contacts.json laden
    dl_contacts = []
    for agent in AGENTS:
        cpath = os.path.join(DATALAKE, agent, "memory", "contacts.json")
        if not os.path.exists(cpath):
            continue
        try:
            with open(cpath) as f:
                data = json.load(f)
            for c in data.get("contacts", []):
                email = (c.get("email") or "").lower().strip()
                if email:
                    c["_agent"] = agent
                    dl_contacts.append(c)
        except Exception as e:
            print(f"  WARN: {cpath}: {e}")

    print(f"  {len(dl_contacts)} Kontakte aus Datalake geladen")

    # Target-DB: die groesste Quelle (A42FFC88)
    target_db = os.path.join(
        ADDRESSBOOK_DIR,
        "Sources/A42FFC88-6123-452D-8D58-9CFE3B556EF6/AddressBook-v22.abcddb",
    )
    if not os.path.exists(target_db):
        # Fallback: erste DB mit E-Mails
        dbs = get_all_dbs()
        if dbs:
            target_db = dbs[0]
        else:
            print("  FEHLER: Keine AddressBook-DB gefunden")
            return 0

    # Globaler E-Mail-Index (alle DBs)
    all_emails = set()
    for db in get_all_dbs():
        idx = build_email_index(db)
        all_emails.update(idx.keys())
    print(f"  {len(all_emails)} E-Mails bereits im Adressbuch")

    # Neue Kontakte identifizieren
    to_add = []
    for c in dl_contacts:
        email = c.get("email", "").lower().strip()
        if email in all_emails:
            continue
        to_add.append(c)

    print(f"  {len(to_add)} neue Kontakte zum Hinzufuegen")

    if not to_add:
        print("  Keine neuen Kontakte — alles schon im Adressbuch.")
        return 0

    # In Target-DB einfuegen
    conn = sqlite3.connect(target_db)
    cur = conn.cursor()

    # Max PKs holen
    cur.execute("SELECT MAX(Z_PK) FROM ZABCDRECORD")
    max_record_pk = cur.fetchone()[0] or 0
    cur.execute("SELECT MAX(Z_PK) FROM ZABCDEMAILADDRESS")
    max_email_pk = cur.fetchone()[0] or 0
    cur.execute("SELECT MAX(Z_PK) FROM ZABCDPHONENUMBER")
    max_phone_pk = cur.fetchone()[0] or 0

    added = 0
    for c in to_add:
        email = c["email"]
        name = (c.get("name") or "").strip()
        company = (c.get("company") or "").strip()
        title = (c.get("title") or "").strip() or None
        phone = (c.get("phone") or "").strip() or None

        fn, ln = split_name(name) if name else ("", email)
        now = datetime.datetime.now().timestamp() - 978307200  # CoreData epoch
        year = datetime.datetime.now().year
        uid = str(uuid.uuid4()).upper()

        # Record einfuegen
        max_record_pk += 1
        rpk = max_record_pk
        cur.execute("""
            INSERT INTO ZABCDRECORD (
                Z_PK, Z_ENT, Z_OPT, ZFIRSTNAME, ZLASTNAME,
                ZORGANIZATION, ZJOBTITLE,
                ZCREATIONDATE, ZMODIFICATIONDATE,
                ZCREATIONDATEYEAR, ZMODIFICATIONDATEYEAR,
                ZCREATIONDATEYEARLESS, ZMODIFICATIONDATEYEARLESS,
                ZUNIQUEID
            ) VALUES (?, ?, 1, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            rpk, Z_ENT_RECORD, fn or None, ln or None,
            company or None, title,
            now, now, year, year, 0.0, 0.0, uid,
        ))

        # E-Mail einfuegen
        max_email_pk += 1
        epk = max_email_pk
        cur.execute("""
            INSERT INTO ZABCDEMAILADDRESS (
                Z_PK, Z_ENT, Z_OPT, ZISPRIMARY, ZORDERINGINDEX,
                ZOWNER, Z22_OWNER,
                ZADDRESS, ZADDRESSNORMALIZED,
                ZLABEL, ZUNIQUEID
            ) VALUES (?, ?, 1, 1, 0, ?, ?, ?, ?, '', ?)
        """, (
            epk, Z_ENT_EMAIL, rpk, Z_ENT_RECORD,
            email, email.lower(),
            str(uuid.uuid4()).upper(),
        ))

        # Telefonnummer falls vorhanden
        if phone:
            max_phone_pk += 1
            ppk = max_phone_pk
            # Letzte 4 Ziffern
            digits_only = re.sub(r"\D", "", phone)
            last4 = digits_only[-4:] if len(digits_only) >= 4 else digits_only
            cur.execute("""
                INSERT INTO ZABCDPHONENUMBER (
                    Z_PK, Z_ENT, Z_OPT, ZISPRIMARY, ZORDERINGINDEX,
                    ZOWNER, Z22_OWNER,
                    ZFULLNUMBER, ZLASTFOURDIGITS,
                    ZLABEL, ZUNIQUEID
                ) VALUES (?, ?, 1, 1, 0, ?, ?, ?, ?, '', ?)
            """, (
                ppk, Z_ENT_PHONE, rpk, Z_ENT_RECORD,
                phone, last4,
                str(uuid.uuid4()).upper(),
            ))

        added += 1

    conn.commit()
    conn.close()
    print(f"  {added} Kontakte ins Adressbuch eingefuegt (DB: {os.path.basename(os.path.dirname(target_db))[:12]})")
    return added


# ═══════════════════════════════════════════════════════════════════════════
# Main
# ═══════════════════════════════════════════════════════════════════════════

def main():
    print()
    print("  APPLE CONTACTS: AUTO-KORREKTUREN + DATALAKE MERGE")
    print("  " + "─" * 50)
    print()

    # Contacts.app beenden (damit DB nicht locked ist)
    print("  Beende Contacts.app falls offen...")
    os.system("osascript -e 'tell application \"Contacts\" to quit' 2>/dev/null")
    import time
    time.sleep(2)

    n_fixes = apply_display_name_fixes()
    n_merged = merge_datalake_contacts()

    print("\n" + "=" * 55)
    print(f"  FERTIG")
    print(f"  Display-Name Korrekturen: {n_fixes}")
    print(f"  Neue Kontakte gemerged:   {n_merged}")
    print("=" * 55)

    # Contacts.app neustarten damit Aenderungen sichtbar werden
    print("\n  Starte Contacts.app neu...")
    os.system("open -a Contacts")
    print("  Fertig. Bitte in Contacts.app pruefen.")
    print(f"\n  Backup: ~/AssistantDev/backups/addressbook_*/")


if __name__ == "__main__":
    main()
