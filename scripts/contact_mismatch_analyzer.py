#!/usr/bin/env python3
"""
Kontakt-Mismatch-Analyse (Vollversion)

Analysiert ALLE Kontakte aus:
  1. macOS Kontaktbuch (AddressBook SQLite-Datenbanken, alle Quellen)
  2. contacts.json Dateien im Datalake
  3. E-Mail-Header (Von:-Zeilen) aus email_*.txt Dateien

Fuer jeden Kontakt mit E-Mail-Adresse wird geprueft ob der gespeicherte
Display-Name zum lokalen E-Mail-Teil passt. Mismatches, Duplikate und
fehlende Namen werden identifiziert.

Aufruf:  python3 ~/AssistantDev/scripts/contact_mismatch_analyzer.py
Output:  ~/Library/.../claude_outputs/contact_mismatch_YYYY-MM-DD.xlsx
"""

import os
import re
import json
import glob
import sqlite3
import datetime
import unicodedata
from collections import defaultdict

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("FEHLER: openpyxl fehlt. pip3 install openpyxl")
    raise SystemExit(1)

# ── Pfade ────────────────────────────────────────────────────────────────────

DATALAKE = os.path.expanduser(
    "~/Library/Mobile Documents/com~apple~CloudDocs/Downloads shared/claude_datalake"
)
OUTPUT_DIR = os.path.expanduser(
    "~/Library/Mobile Documents/com~apple~CloudDocs/Downloads shared/claude_outputs"
)
ADDRESSBOOK_DIR = os.path.expanduser("~/Library/Application Support/AddressBook")
AGENTS = ["signicat", "trustedcarrier", "privat", "standard", "system ward"]


# ═══════════════════════════════════════════════════════════════════════════
# Hilfsfunktionen
# ═══════════════════════════════════════════════════════════════════════════

def normalize(s):
    """Normalisiert: lowercase, diakritische Zeichen entfernen, nur a-z/Leerzeichen.
    Deutsche Umlaut-Aequivalente (ae/oe/ue) werden auf die Basis-Vokale reduziert,
    sodass 'Maennel'='Männel' und 'Moeller'='Möller' als gleich erkannt werden.
    """
    if not s:
        return ""
    s = unicodedata.normalize("NFD", s)
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    s = s.lower().strip()
    # Deutsche Umlaut-Schreibweisen auf Basis-Vokal normalisieren:
    # ae->a, oe->o, ue->u (NACH NFD, das bereits ä->a etc. gemacht hat)
    s = s.replace("ae", "a").replace("oe", "o").replace("ue", "u")
    s = re.sub(r"[^a-z\s]", " ", s)
    return re.sub(r"\s+", " ", s).strip()


def tokens(s):
    return {t for t in normalize(s).split() if len(t) >= 2}


GENERIC_LOCAL_PARTS = {
    "info", "support", "no-reply", "noreply", "admin", "contact", "office",
    "team", "hello", "help", "sales", "billing", "newsletter", "marketing",
    "service", "notification", "notifications", "notificacao", "alerts",
    "alert", "mailer", "postmaster", "webmaster", "do-not-reply", "donotreply",
    "invoice", "invoices", "hr", "careers", "jobs", "press", "contato",
    "atendimento", "sac", "faleconosco", "cadastro", "comunicacao",
    "financeiro", "comercial", "mail", "email", "messages", "updates",
}


def _looks_like_real_name(name_str, email_addr=""):
    """Prueft ob ein extrahierter Name plausibel nach einem echten Namen aussieht.
    Filtert:
    - Hash-artige E-Mail-Prefixe (Booking-IDs, VTEX, UUID-artige Strings)
    - Tokens ohne Vokale oder zu kurz
    - Prefixe die mit langen Nummern beginnen
    """
    if not name_str:
        return False
    # Check 1: E-Mail-Prefix direkt — Hash/ID-Erkennung
    if email_addr and "@" in email_addr:
        local = email_addr.split("@")[0]
        # Startet mit ≥4 Ziffern → Booking-ID, Tracking-Nummer etc.
        if re.match(r"^\d{4,}", local):
            return False
        # Enthaelt UUID/Hex-Muster (≥8 hex chars am Stueck)
        if re.search(r"[0-9a-f]{8,}", local.lower()):
            return False
        # Mehr Ziffern als Buchstaben → kein Name
        digits = sum(1 for c in local if c.isdigit())
        alpha = sum(1 for c in local if c.isalpha())
        if digits > alpha and digits >= 3:
            return False

    # Check 2: Laenge des lokalen Teils — Hashes/Tokens sind typisch >25 Zeichen
    if email_addr and "@" in email_addr:
        local = email_addr.split("@")[0]
        # Einzelnes Wort ohne Trenner das laenger als 18 Zeichen ist → Hash/Token
        unsep = re.sub(r"[._-]", "", local)
        if len(unsep) > 18 and not re.search(r"[._-]", local):
            return False
        # Auch mit Trennern: wenn der laengste Teil >15 Zeichen ist → suspekt
        parts_raw = re.split(r"[._-]", local)
        if any(len(p) > 15 for p in parts_raw):
            return False

    # Check 3: Namens-Tokens pruefen — mindestens 2 plausible (Vorname + Nachname)
    parts = name_str.split()
    vowels = set("aeiouAEIOU")
    plausible = 0
    for p in parts:
        clean = p.replace("-", "").replace("'", "")
        if len(clean) < 2:
            continue
        if not clean.isalpha():
            continue
        if not (vowels & set(clean)):
            continue
        # Namens-Token: 3-20 Zeichen, hat Vokal, rein alphabetisch
        if 3 <= len(clean) <= 20:
            plausible += 1
    # Mindestens 2 plausible Tokens (Vorname + Nachname)
    return plausible >= 2


def _strip_trailing_digits(s):
    """Entfernt Zahlen am Ende: 'annelang1982' -> 'annelang', 'ali0503' -> 'ali'."""
    return re.sub(r"\d+$", "", s)


def name_from_email(email_addr):
    """Versucht einen Klarnamen aus dem E-Mail-Prefix zu extrahieren.
    Returns (name_str, confidence: 'high' | 'low' | 'unclear').

    Verbesserte Logik:
    - Trailing Digits werden ignoriert (Geburtsjahre, IDs: 1985, 001, 23)
    - Punkte/Bindestriche/Unterstriche sind Wort-Trenner
    - Einzelbuchstaben vor einem Nachnamen werden als Initial akzeptiert
    - camelCase-Erkennung
    - Concatenated Names (annelang -> 'low' confidence fuer spaeteres Matching)
    """
    if not email_addr or "@" not in email_addr:
        return "", "unclear"
    local = email_addr.split("@")[0].lower()

    # Schritt 1: Generische Adressen filtern
    clean = local.replace("-", "").replace("_", "").replace(".", "")
    if clean in GENERIC_LOCAL_PARTS or local in GENERIC_LOCAL_PARTS:
        return "", "unclear"
    # Rein numerisch (z.B. 017625817887)
    if re.match(r"^\d+$", local):
        return "", "unclear"
    # Nur 1-2 Zeichen ohne Punkt/Trenner und keine Digits → zu wenig Info
    if len(local) < 2:
        return "", "unclear"

    # Schritt 2: Trailing Digits entfernen (1982, 001, 29, etc.)
    stripped = _strip_trailing_digits(local)
    if not stripped:
        return "", "unclear"

    # Schritt 3: An Trennzeichen splitten
    parts = re.split(r"[._-]", stripped)
    parts = [p for p in parts if p and not re.match(r"^\d+$", p)]

    if len(parts) >= 2:
        # z.B. a.roman -> ["a", "roman"], vorname.nachname -> ["vorname", "nachname"]
        name = " ".join(p.capitalize() for p in parts)
        # Wenn erster Teil nur 1 Buchstabe (Initial): trotzdem 'high' confidence
        return name, "high"

    if len(parts) == 1:
        part = parts[0]

        # camelCase: "philippWegmann" → "Philipp Wegmann"
        camel_split = re.findall(r"[A-Z]?[a-z]+|[A-Z]+(?=[A-Z][a-z]|$)", part)
        if len(camel_split) >= 2:
            return " ".join(w.capitalize() for w in camel_split), "high"

        # 2 Buchstaben = moegliche Initialen (z.B. "ab", "ad", "ah")
        if len(part) == 2 and part.isalpha():
            return part.upper(), "initials"

        # 3 Buchstaben = moeglicherweise Initial + Kurzname (z.B. "ala" = A.La)
        if len(part) == 3 and part.isalpha():
            return part.capitalize(), "low"

        # 4+ Buchstaben = koennte zusammengeschriebener Name sein
        if len(part) >= 4 and part.isalpha():
            return part.capitalize(), "low"

    return "", "unclear"


def compare_names(stored_name, extracted_name, confidence="high"):
    """Vergleicht zwei Namen mit erweiterter Logik.
    Returns ('OK' | 'MISMATCH' | 'UNKLAR', note_str).

    Verbesserte Checks:
    1. Exakte Token-Uebereinstimmung
    2. Teilstring-Match (ein Token ist Substring des anderen)
    3. Initial-Matching: 'A' aus E-Mail matcht 'andreas' im gespeicherten Namen
    4. Concatenated-Name-Splitting: 'annelang' gegen ['anne','lang'] pruefen
    """
    if not stored_name or not stored_name.strip():
        return "UNKLAR", "Kein Display-Name gespeichert"
    if not extracted_name:
        return "UNKLAR", "Name aus E-Mail nicht extrahierbar"

    stored_tok = tokens(stored_name)
    extracted_tok = tokens(extracted_name)
    if not stored_tok or not extracted_tok:
        # Nicht-lateinische Zeichen (CJK, Kyrillisch) oder Kurzformen → OK mit Hinweis
        return "OK", "Nicht-lateinischer Name oder Kurzform — automatisch akzeptiert"

    # Check 1: Direkte Token-Uebereinstimmung
    overlap = stored_tok & extracted_tok
    if overlap:
        return "OK", f"Match: {', '.join(sorted(overlap))}"

    # Check 2: Teilstring-Match (z.B. 'philipp' in 'phiweg' oder umgekehrt)
    for st in stored_tok:
        for et in extracted_tok:
            if len(st) >= 3 and len(et) >= 3 and (st in et or et in st):
                return "OK", f"Teilmatch: '{st}' ~ '{et}'"

    # Check 3: Initial-Matching
    # z.B. extracted_tok = {'a', 'roman'} vs. stored_tok = {'astrid', 'roman'}
    # → 'a' ist 1 Buchstabe → pruefe ob es der Anfangsbuchstabe eines stored token ist
    extracted_initials = {t for t in extracted_tok if len(t) == 1}
    extracted_words = {t for t in extracted_tok if len(t) >= 2}
    stored_initials_set = {t[0] for t in stored_tok if t}  # erste Buchstaben aller stored tokens

    if extracted_initials:
        # Mindestens ein Wort-Token muss matchen UND die Initialen muessen passen
        word_match = extracted_words & stored_tok
        init_match = extracted_initials & stored_initials_set
        if word_match and init_match:
            return "OK", f"Initial-Match: {sorted(init_match)} + Wort: {sorted(word_match)}"
        # Nur Initialen (2 Buchstaben, z.B. "AB"): beide muessen passen
        if len(extracted_initials) >= 2 and not extracted_words:
            if extracted_initials <= stored_initials_set:
                return "OK", f"Initialen-Match: {sorted(extracted_initials)} ~ {sorted(stored_tok)}"

    # Check 3b: Einzelnes Initial + 1 Wort, Wort matcht Nachname mit Teilstring
    if len(extracted_initials) == 1 and len(extracted_words) == 1:
        ew = list(extracted_words)[0]
        ei = list(extracted_initials)[0]
        for st in stored_tok:
            if len(st) >= 3 and len(ew) >= 3 and (st in ew or ew in st):
                if ei in stored_initials_set:
                    return "OK", f"Initial '{ei}' + Teilmatch '{ew}' ~ '{st}'"

    # Check 4: Concatenated Name Splitting
    # z.B. extracted = "annelang" (low confidence) → pruefe ob stored tokens
    # als Substrings zusammenhaengend drin vorkommen
    if confidence == "low" and len(extracted_tok) == 1:
        concat = list(extracted_tok)[0]  # z.B. "annelang", "amuminovic"
        if len(concat) >= 5:
            # Pruefe ob 2 stored tokens zusammengesetzt den concat ergeben
            stored_list = sorted(stored_tok, key=len, reverse=True)
            for st in stored_list:
                if len(st) >= 3 and st in concat:
                    remaining = concat.replace(st, "", 1)
                    if remaining:
                        for st2 in stored_list:
                            if st2 != st and len(st2) >= 2 and (
                                remaining.startswith(st2[:3]) or
                                st2.startswith(remaining[:3])
                            ):
                                return "OK", f"Concat-Match: '{concat}' enthaelt '{st}' + '{st2}'"
                    # Auch nur ein Match im concat ist schon ein gutes Zeichen
                    rest = concat.replace(st, "", 1)
                    # Pruefe ob der Rest ein Initial eines anderen stored tokens ist
                    if len(rest) <= 2 and rest:
                        if rest[0] in stored_initials_set:
                            return "OK", f"Concat-Match: '{concat}' = '{rest[0]}'+'{st}'"

    # Check 5: Confidence 'initials' — 2-Buchstaben-Local-Part
    # z.B. "AB" extracted → pruefe ob A und B die Initialen von Vor+Nachname sind
    if confidence == "initials" and len(extracted_tok) == 1:
        init_str = normalize(extracted_name)  # z.B. "ab"
        if len(init_str) == 2:
            stored_sorted = sorted(stored_tok)
            # Pruefe alle Permutationen von 2 stored tokens
            for i, s1 in enumerate(stored_sorted):
                for s2 in stored_sorted[i+1:]:
                    if {s1[0], s2[0]} == set(init_str):
                        return "OK", f"Initialen '{init_str}' = '{s1[0]}({s1})' + '{s2[0]}({s2})'"
            # Auch Reihenfolge-unabhaengig
            if set(init_str) <= stored_initials_set:
                return "OK", f"Initialen '{init_str}' matchen Anfangsbuchstaben von {sorted(stored_tok)}"

    return "MISMATCH", (
        f"Keine Uebereinstimmung: gespeichert={sorted(stored_tok)} "
        f"vs. email={sorted(extracted_tok)}"
    )


# ═══════════════════════════════════════════════════════════════════════════
# Datenquellen: macOS Contacts (AddressBook SQLite)
# ═══════════════════════════════════════════════════════════════════════════

def load_macos_contacts():
    """Liest alle Kontakte mit E-Mail aus allen AddressBook-Quellen.
    Returns list of dicts.
    """
    dbs = glob.glob(os.path.join(ADDRESSBOOK_DIR, "Sources/*/AddressBook-v22.abcddb"))
    # Haupt-DB auch pruefen
    main_db = os.path.join(ADDRESSBOOK_DIR, "AddressBook-v22.abcddb")
    if os.path.exists(main_db):
        dbs.append(main_db)

    contacts = []
    seen_emails = set()  # Deduplizierung ueber Quellen
    for db_path in dbs:
        src_id = os.path.basename(os.path.dirname(db_path))[:12]
        try:
            conn = sqlite3.connect(db_path)
            cur = conn.cursor()
            cur.execute("""
                SELECT r.ZFIRSTNAME, r.ZLASTNAME, r.ZMIDDLENAME, r.ZNICKNAME,
                       r.ZORGANIZATION, e.ZADDRESS, r.Z_PK
                FROM ZABCDRECORD r
                LEFT JOIN ZABCDEMAILADDRESS e ON e.ZOWNER = r.Z_PK
                WHERE e.ZADDRESS IS NOT NULL AND e.ZADDRESS != ''
            """)
            for row in cur.fetchall():
                firstname = (row[0] or "").strip()
                lastname = (row[1] or "").strip()
                middlename = (row[2] or "").strip()
                nickname = (row[3] or "").strip()
                org = (row[4] or "").strip()
                email = (row[5] or "").strip().lower()
                if not email or "@" not in email:
                    continue
                if email in seen_emails:
                    continue
                seen_emails.add(email)

                # Display-Name zusammenbauen
                name_parts = [p for p in [firstname, middlename, lastname] if p]
                display_name = " ".join(name_parts)
                # Wenn Nachname = E-Mail-Adresse (Bug/Artefakt): Name leer
                if display_name.lower() == email or "@" in display_name:
                    display_name = ""

                contacts.append({
                    "source": f"macOS ({src_id})",
                    "display_name": display_name,
                    "firstname": firstname,
                    "lastname": lastname,
                    "nickname": nickname,
                    "organization": org,
                    "email": email,
                })
            conn.close()
        except Exception as e:
            print(f"  WARN: DB {src_id} Fehler: {e}")
    return contacts


# ═══════════════════════════════════════════════════════════════════════════
# Datenquellen: Datalake contacts.json
# ═══════════════════════════════════════════════════════════════════════════

def load_datalake_contacts():
    contacts = []
    for agent in AGENTS:
        cpath = os.path.join(DATALAKE, agent, "memory", "contacts.json")
        if not os.path.exists(cpath):
            continue
        try:
            with open(cpath, "r", encoding="utf-8") as f:
                data = json.load(f)
            for c in data.get("contacts", []):
                email = (c.get("email") or "").lower().strip()
                if not email:
                    continue
                contacts.append({
                    "source": f"datalake ({agent})",
                    "display_name": (c.get("name") or "").strip(),
                    "email": email,
                    "last_contact": c.get("last_contact", ""),
                    "organization": c.get("company", ""),
                })
        except Exception as e:
            print(f"  WARN: {cpath}: {e}")
    return contacts


# ═══════════════════════════════════════════════════════════════════════════
# Datenquellen: E-Mail-Header (Von:-Zeilen)
# ═══════════════════════════════════════════════════════════════════════════

EMAIL_FROM_RE = re.compile(r"^Von:\s*(.+)$", re.MULTILINE | re.IGNORECASE)
DISPLAY_FROM_RE = re.compile(r"^\s*(.+?)\s*<([\w.+%-]+@[\w.-]+\.[a-zA-Z]{2,})>")
EMAIL_ADDR_RE = re.compile(r"[\w.+%-]+@[\w.-]+\.[a-zA-Z]{2,}")
EMAIL_FILE_RE = re.compile(r"^\d{4}-\d{2}-\d{2}_\d{2}-\d{2}-\d{2}_(IN|OUT)_.*\.txt$")
LEGACY_EMAIL_RE = re.compile(r"^email_.*\.txt$", re.IGNORECASE)


def load_email_header_names():
    """Scant alle E-Mail-.txt Dateien im Datalake und gibt
    {email_lower: set(display_names)} zurueck.
    """
    result = defaultdict(set)
    for agent in AGENTS:
        memdir = os.path.join(DATALAKE, agent, "memory")
        if not os.path.isdir(memdir):
            continue
        try:
            entries = os.listdir(memdir)
        except Exception:
            continue
        for fname in entries:
            if not (EMAIL_FILE_RE.match(fname) or LEGACY_EMAIL_RE.match(fname)):
                continue
            fpath = os.path.join(memdir, fname)
            try:
                with open(fpath, "r", encoding="utf-8", errors="replace") as f:
                    head = f.read(2048)
            except Exception:
                continue
            m = EMAIL_FROM_RE.search(head)
            if not m:
                continue
            dm = DISPLAY_FROM_RE.match(m.group(1).strip())
            if dm:
                display = dm.group(1).strip().strip('"').strip("'")
                addr = dm.group(2).lower()
                if display and "@" not in display:
                    result[addr].add(display)
    return result


# ═══════════════════════════════════════════════════════════════════════════
# Analyse
# ═══════════════════════════════════════════════════════════════════════════

def analyze():
    print("=" * 60)
    print("  KONTAKT-MISMATCH-ANALYSE (macOS Contacts + Datalake)")
    print("=" * 60)
    print()

    # 1) Laden
    print("1) macOS Kontaktbuch laden...")
    macos_contacts = load_macos_contacts()
    print(f"   {len(macos_contacts)} Kontakte mit E-Mail aus macOS AddressBook")

    print("2) Datalake contacts.json laden...")
    dl_contacts = load_datalake_contacts()
    print(f"   {len(dl_contacts)} Kontakte aus Datalake")

    print("3) E-Mail-Header (Von:-Zeilen) scannen...")
    email_headers = load_email_header_names()
    print(f"   {len(email_headers)} unique Absender-Adressen mit Display-Name")
    print()

    # 2) Zusammenfuehren: eine Zeile pro E-Mail-Adresse
    # Prioritaet: macOS Contacts > Datalake contacts.json > E-Mail-Header
    merged = {}  # email -> dict

    for c in macos_contacts:
        email = c["email"]
        if email not in merged:
            merged[email] = {
                "email": email,
                "source": c["source"],
                "display_name": c["display_name"],
                "organization": c.get("organization", ""),
                "last_contact": "",
                "all_sources": [],
            }
        merged[email]["all_sources"].append(c["source"])
        if not merged[email]["display_name"] and c["display_name"]:
            merged[email]["display_name"] = c["display_name"]

    for c in dl_contacts:
        email = c["email"]
        if email not in merged:
            merged[email] = {
                "email": email,
                "source": c["source"],
                "display_name": c["display_name"],
                "organization": c.get("organization", ""),
                "last_contact": c.get("last_contact", ""),
                "all_sources": [],
            }
        merged[email]["all_sources"].append(c["source"])
        if not merged[email]["display_name"] and c["display_name"]:
            merged[email]["display_name"] = c["display_name"]
        if c.get("last_contact"):
            merged[email]["last_contact"] = c["last_contact"]

    print(f"4) {len(merged)} unique E-Mail-Adressen nach Zusammenfuehrung")

    # 3) Analyse
    results_mismatch = []
    results_unclear = []
    results_ok = []
    domain_mismatches = defaultdict(int)

    for email, entry in sorted(merged.items()):
        stored_name = entry["display_name"]
        extracted, confidence = name_from_email(email)
        domain = email.split("@")[1] if "@" in email else ""

        # E-Mail-Header Cross-Check
        header_names = email_headers.get(email, set())
        header_mismatch_note = ""
        if header_names and stored_name:
            for hn in header_names:
                if tokens(hn) and tokens(stored_name) and not (tokens(hn) & tokens(stored_name)):
                    header_mismatch_note = f"E-Mail-Header '{hn}' weicht ab von '{stored_name}'"
                    break

        # UNKLAR-Aufloesung: wenn kein Display-Name gespeichert oder E-Mail
        # nicht extrahierbar → E-Mail-Adresse WIRD zum Display-Namen.
        # Damit gibt es kein "UNKLAR" mehr — alles wird entweder OK oder
        # bekommt die E-Mail als Display-Name und wird als OK gewertet.
        effective_name = stored_name
        if not effective_name:
            effective_name = email
            entry["display_name"] = email  # in merged-Entry uebernehmen
        if confidence == "unclear":
            # Generische Adresse, Name nicht extrahierbar → E-Mail als Name, OK
            status = "OK"
            note = f"Display-Name aus E-Mail-Adresse uebernommen"
        elif (confidence in ("low", "initials")) and not stored_name:
            status = "OK"
            note = f"Display-Name aus E-Mail-Adresse uebernommen (Token: {extracted})"
        else:
            status, note = compare_names(effective_name, extracted, confidence=confidence)

        # Header-Mismatch kann Status ueberschreiben
        if header_mismatch_note and status == "OK":
            status = "MISMATCH"
            note = header_mismatch_note

        row = {
            "source": "; ".join(sorted(set(entry["all_sources"]))) or entry["source"],
            "display_name": stored_name,
            "email": email,
            "extracted_name": extracted,
            "organization": entry.get("organization", ""),
            "domain": domain,
            "last_contact": entry.get("last_contact", ""),
            "email_header_names": "; ".join(sorted(header_names)),
            "note": (note + (" / " + header_mismatch_note if header_mismatch_note and header_mismatch_note not in note else "")).strip(" /"),
        }

        # Auto-Korrektur bei MISMATCH:
        # - Klare Vor.Nachname-Struktur → extrahierten Namen uebernehmen
        # - Sonst → E-Mail-Adresse als Display-Name
        # - Domain wird als Organisation gesetzt
        if status == "MISMATCH":
            domain_mismatches[domain] += 1
            old_name = row["display_name"]
            if confidence == "high" and extracted and _looks_like_real_name(extracted, email):
                # Klare Struktur: Vorname Nachname aus E-Mail
                row["display_name"] = extracted
                row["note"] = f"AUTO-KORRIGIERT: '{old_name}' → '{extracted}' (aus E-Mail extrahiert)"
            else:
                # Keine klare Struktur oder Zufalls-Tokens → E-Mail als Display-Name
                row["display_name"] = email
                row["note"] = f"AUTO-KORRIGIERT: '{old_name}' → E-Mail als Display-Name"
            if domain and not row.get("organization"):
                # Domain als Organisation: 'signicat.com' → 'Signicat'
                dom_parts = domain.split(".")
                row["organization"] = dom_parts[-2].capitalize() if len(dom_parts) >= 2 else domain
            results_mismatch.append(row)
        elif status == "UNKLAR":
            results_unclear.append(row)
        else:
            results_ok.append(row)

    total = len(merged)
    print(f"\n{'=' * 50}")
    print(f"ERGEBNIS: {total} Kontakte analysiert")
    print(f"  OK:         {len(results_ok):>5}")
    print(f"  MISMATCH:   {len(results_mismatch):>5}")
    print(f"  UNKLAR:     {len(results_unclear):>5}")
    print(f"{'=' * 50}")

    # ═══════════════════════════════════════════════════════════════════════
    # Excel erstellen
    # ═══════════════════════════════════════════════════════════════════════
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    today = datetime.datetime.now().strftime("%Y-%m-%d")
    xlsx_path = os.path.join(OUTPUT_DIR, f"contact_mismatch_{today}.xlsx")

    wb = Workbook()
    hfont = Font(bold=True, color="FFFFFF", size=11)
    fill_red = PatternFill(start_color="C0392B", end_color="C0392B", fill_type="solid")
    fill_orange = PatternFill(start_color="E67E22", end_color="E67E22", fill_type="solid")
    fill_green = PatternFill(start_color="27AE60", end_color="27AE60", fill_type="solid")
    fill_blue = PatternFill(start_color="2980B9", end_color="2980B9", fill_type="solid")
    thin = Border(*(Side(style="thin", color="DDDDDD"),) * 4)
    wrap = Alignment(wrap_text=True, vertical="top")

    def write_header(ws, cols, fill):
        for i, c in enumerate(cols, 1):
            cell = ws.cell(row=1, column=i, value=c)
            cell.font = hfont
            cell.fill = fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin

    def style_data(ws, ncols):
        for row in ws.iter_rows(min_row=2, max_col=ncols):
            for cell in row:
                cell.border = thin
                cell.alignment = wrap

    # Sheet 1: AUTO-KORREKTUREN (ehemals KLARER_MISMATCH)
    ws1 = wb.active
    ws1.title = "AUTO_KORREKTUREN"
    c1 = ["Quelle", "Neuer Display-Name", "E-Mail-Adresse", "Name aus E-Mail", "Organisation", "E-Mail-Header-Name", "Letzter Kontakt", "Korrektur-Notiz"]
    write_header(ws1, c1, fill_red)
    results_mismatch.sort(key=lambda r: r["email"])
    for r in results_mismatch:
        ws1.append([r["source"], r["display_name"], r["email"], r["extracted_name"], r["organization"], r["email_header_names"], r["last_contact"], r["note"]])
    style_data(ws1, len(c1))
    for i, w in enumerate([16, 30, 38, 24, 22, 28, 14, 55], 1):
        ws1.column_dimensions[chr(64 + i)].width = w

    # Sheet 2: UNKLAR
    ws2 = wb.create_sheet("UNKLAR")
    c2 = ["Quelle", "Display-Name", "E-Mail-Adresse", "Organisation", "Grund", "Letzter Kontakt"]
    write_header(ws2, c2, fill_orange)
    results_unclear.sort(key=lambda r: r["email"])
    for r in results_unclear:
        ws2.append([r["source"], r["display_name"], r["email"], r["organization"], r["note"], r["last_contact"]])
    style_data(ws2, len(c2))
    for i, w in enumerate([16, 28, 38, 22, 55, 14], 1):
        ws2.column_dimensions[chr(64 + i)].width = w

    # Sheet 3: ZUSAMMENFASSUNG
    ws3 = wb.create_sheet("ZUSAMMENFASSUNG")
    c3 = ["Metrik", "Wert"]
    write_header(ws3, c3, fill_blue)
    ws3.append(["Analysedatum", today])
    ws3.append(["Datenquellen", "macOS Contacts + Datalake contacts.json + E-Mail-Header"])
    ws3.append(["macOS Kontakte mit E-Mail", len(macos_contacts)])
    ws3.append(["Datalake Kontakte", len(dl_contacts)])
    ws3.append(["E-Mail-Header Absender", len(email_headers)])
    ws3.append(["Unique E-Mails (merged)", total])
    ws3.append([])
    ws3.append(["Ergebnis OK", len(results_ok)])
    ws3.append(["Ergebnis MISMATCH", len(results_mismatch)])
    ws3.append(["Ergebnis UNKLAR", len(results_unclear)])
    ws3.append([])
    ws3.append(["Top-Domains mit Mismatches", ""])
    for dom, cnt in sorted(domain_mismatches.items(), key=lambda x: x[1], reverse=True)[:20]:
        ws3.append([f"  @{dom}", cnt])
    style_data(ws3, 2)
    ws3.column_dimensions["A"].width = 38
    ws3.column_dimensions["B"].width = 22

    # Sheet 4: ALLE_OK (kompakt)
    ws4 = wb.create_sheet("ALLE_OK")
    c4 = ["E-Mail-Adresse", "Display-Name", "Organisation", "Quelle"]
    write_header(ws4, c4, fill_green)
    results_ok.sort(key=lambda r: r["email"])
    for r in results_ok:
        ws4.append([r["email"], r["display_name"], r["organization"], r["source"]])
    style_data(ws4, len(c4))
    for i, w in enumerate([38, 30, 22, 16], 1):
        ws4.column_dimensions[chr(64 + i)].width = w

    wb.save(xlsx_path)
    print(f"\nExcel gespeichert: {xlsx_path}")
    print(f"  Sheet AUTO_KORREKTUREN: {len(results_mismatch)} Eintraege (Display-Name auto-korrigiert)")
    print(f"  Sheet UNKLAR:           {len(results_unclear)} Eintraege")
    print(f"  Sheet ALLE_OK:          {len(results_ok)} Eintraege")
    print(f"  Sheet ZUSAMMENFASSUNG:  Statistik + Top-Domains")

    return xlsx_path


if __name__ == "__main__":
    analyze()
