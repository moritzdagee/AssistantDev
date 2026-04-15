#!/usr/bin/env python3
"""
extract_contacts.py — Zweistufige Kontakt-Extraktion aus E-Mails.

STUFE 1: Alle E-Mails scannen (kein API-Call)
  - From-Header parsen → Name + Email
  - Automatisierte Absender rausfiltern
  - Duplikate zusammenfuehren (laengsten Body merken)
  → Vorschau-Excel mit Name + Email

STUFE 2: Signatur-Extraktion via Claude Haiku API (nur eindeutige Kontakte)
  - Letzte 60 Zeilen als Signatur-Kandidat
  - Batches von 20 Signaturen pro API-Call
  - Checkpoint-Datei fuer Fortsetzbarkeit (--continue)
  → Finales Excel + vCard

Ausfuehren:
  python3 ~/AssistantDev/scripts/extract_contacts.py
  python3 ~/AssistantDev/scripts/extract_contacts.py --continue   # Stufe 2 fortsetzen
"""

import os
import sys
import json
import email
import email.header
import email.utils
import re
import time
import datetime
import subprocess

# ─── PATHS ────────────────────────────────────────────────────────────────────

BASE = os.path.expanduser(
    "~/Library/Mobile Documents/com~apple~CloudDocs/Downloads shared/claude_datalake"
)
INBOX_DIR = os.path.join(BASE, "email_inbox")
OUTPUT_DIR = os.path.join(BASE, "..", "claude_outputs")
MODELS_FILE = os.path.join(BASE, "config/models.json")
CHECKPOINT_FILE = os.path.expanduser("~/.extract_contacts_progress.json")

# ─── FILTER LISTS ─────────────────────────────────────────────────────────────

SKIP_LOCAL_PARTS = {
    'noreply', 'no-reply', 'newsletter', 'mailer', 'notifications',
    'donotreply', 'support', 'info', 'hello', 'contact', 'sales',
    'marketing', 'postmaster', 'bounce', 'autoresponder', 'daemon',
    'mailer-daemon', 'notify', 'automated', 'do-not-reply',
}

SKIP_DOMAINS = {
    'mailchimp.com', 'sendgrid.net', 'sendgrid.com', 'hubspot.com',
    'salesforce.com', 'linkedin.com', 'google.com', 'facebook.com',
    'twitter.com', 'amazon.com', 'apple.com', 'microsoft.com',
    'github.com', 'slack.com', 'zoom.us', 'zoom.com', 'dropbox.com',
    'mailgun.com', 'mandrillapp.com', 'constantcontact.com',
    'sendinblue.com', 'brevo.com', 'klaviyo.com', 'intercom.io',
    'postmarkapp.com', 'sparkpost.com', 'amazonses.com',
    'googlemail.com',  # Google system mails
}

# Eigene Adressen — Mails von diesen Adressen NIE als fremder Kontakt importieren.
# Schuetzt vor My-Card-Pollution beim vCard-Import.
OWN_EMAILS = {
    'moritz.cremer@me.com', 'moritz.cremer@icloud.com',
    'moritz.cremer@signicat.com', 'londoncityfox@gmail.com',
    'moritz@demoscapital.co', 'moritz@vegatechnology.com.br',
    'moritz.cremer@trustedcarrier.net', 'moritz@brandshare.me',
    'cremer.moritz@gmx.de', 'family.cremer@gmail.com',
    'moritz.cremer@trustedcarrier.de', 'naiaraebertz@gmail.com',
}

HAIKU_COST_PER_1K_INPUT = 0.00025   # $0.25 / 1M input tokens
HAIKU_COST_PER_1K_OUTPUT = 0.00125  # $1.25 / 1M output tokens
EST_TOKENS_PER_BATCH = 2500         # ~2500 input tokens per batch of 20
EST_OUTPUT_PER_BATCH = 800          # ~800 output tokens per batch
BATCH_SIZE = 20

# ─── HELPERS ──────────────────────────────────────────────────────────────────


def decode_header(raw):
    """Decode RFC 2047 encoded header."""
    if not raw:
        return ''
    parts = email.header.decode_header(raw)
    decoded = []
    for part, charset in parts:
        if isinstance(part, bytes):
            decoded.append(part.decode(charset or 'utf-8', errors='replace'))
        else:
            decoded.append(part)
    return ' '.join(decoded).strip()


def get_body_text(msg):
    """Extract plain text body from email.message.Message."""
    if msg.is_multipart():
        for part in msg.walk():
            if part.get_content_type() == 'text/plain':
                payload = part.get_payload(decode=True)
                if payload:
                    charset = part.get_content_charset() or 'utf-8'
                    return payload.decode(charset, errors='replace')
        # Fallback: text/html with tags stripped
        for part in msg.walk():
            if part.get_content_type() == 'text/html':
                payload = part.get_payload(decode=True)
                if payload:
                    charset = part.get_content_charset() or 'utf-8'
                    html = payload.decode(charset, errors='replace')
                    return re.sub(r'<[^>]+>', ' ', html).strip()
    else:
        payload = msg.get_payload(decode=True)
        if payload:
            charset = msg.get_content_charset() or 'utf-8'
            return payload.decode(charset, errors='replace')
    return ''


def _name_tokens(name):
    """Lowercased, umlaut-normalized alpha tokens (>= 2 chars) of a display name."""
    if not name:
        return set()
    s = name.lower()
    for a, b in [('ä', 'ae'), ('ö', 'oe'), ('ü', 'ue'), ('ß', 'ss')]:
        s = s.replace(a, b)
    return {t for t in re.findall(r'[a-z]+', s) if len(t) >= 2}


def name_matches_email(name, email_addr):
    """True if at least one >=3 char name token appears in the email's local-part.
    If the name has no qualifying tokens (length >=3), accept by default — many
    legitimate display names are pure initials or short company tags.
    """
    if not name or '@' not in email_addr:
        return True
    local = email_addr.split('@')[0].lower()
    local_clean = re.sub(r'[._\-+]', ' ', local)
    qual_tokens = [t for t in _name_tokens(name) if len(t) >= 3]
    if not qual_tokens:
        return True
    return any(t in local_clean for t in qual_tokens)


def is_strict_name_extension(old_name, new_name):
    """True if new_name strictly extends old_name (all old tokens contained in new)."""
    if not old_name:
        return True
    old_t = _name_tokens(old_name)
    new_t = _name_tokens(new_name)
    return bool(old_t) and old_t.issubset(new_t)


def is_automated_sender(email_addr):
    """Check if email address belongs to an automated/mass sender."""
    if not email_addr or '@' not in email_addr:
        return True
    local, domain = email_addr.rsplit('@', 1)
    local = local.lower()
    domain = domain.lower()
    # Check local part
    for skip in SKIP_LOCAL_PARTS:
        if skip in local:
            return True
    # Check domain
    for skip_domain in SKIP_DOMAINS:
        if domain == skip_domain or domain.endswith('.' + skip_domain):
            return True
    return False


def load_api_key():
    """Load Anthropic API key from models.json."""
    try:
        with open(MODELS_FILE) as f:
            config = json.load(f)
        return config['providers']['anthropic']['api_key']
    except Exception as e:
        print(f"Fehler beim Laden des API Keys: {e}")
        sys.exit(1)


def load_checkpoint():
    """Load checkpoint file for resuming stage 2."""
    if os.path.exists(CHECKPOINT_FILE):
        try:
            with open(CHECKPOINT_FILE) as f:
                return json.load(f)
        except Exception:
            pass
    return None


def save_checkpoint(data):
    """Save checkpoint file."""
    with open(CHECKPOINT_FILE, 'w') as f:
        json.dump(data, f, indent=2, ensure_ascii=False)


def format_time(seconds):
    """Format seconds into human-readable string."""
    if seconds < 60:
        return f"{int(seconds)}s"
    elif seconds < 3600:
        return f"{int(seconds // 60)}m {int(seconds % 60)}s"
    else:
        return f"{int(seconds // 3600)}h {int((seconds % 3600) // 60)}m"


# ─── vCARD GENERATION ─────────────────────────────────────────────────────────

def contact_to_vcard(c):
    """Convert a contact dict to vCard 3.0 string."""
    lines = ['BEGIN:VCARD', 'VERSION:3.0']
    name = c.get('name', '')
    if name:
        parts = name.split(None, 1)
        if len(parts) == 2:
            lines.append(f'N:{parts[1]};{parts[0]};;;')
        else:
            lines.append(f'N:{name};;;;')
        lines.append(f'FN:{name}')
    if c.get('email'):
        lines.append(f'EMAIL;TYPE=INTERNET:{c["email"]}')
    if c.get('phone'):
        lines.append(f'TEL;TYPE=WORK:{c["phone"]}')
    if c.get('mobile'):
        lines.append(f'TEL;TYPE=CELL:{c["mobile"]}')
    if c.get('company'):
        lines.append(f'ORG:{c["company"]}')
    if c.get('title'):
        lines.append(f'TITLE:{c["title"]}')
    if c.get('website'):
        lines.append(f'URL:{c["website"]}')
    if c.get('address'):
        addr = c['address'].replace('\n', '\\n')
        lines.append(f'ADR;TYPE=WORK:;;{addr};;;;')
    lines.append('END:VCARD')
    return '\n'.join(lines)


# ─── STUFE 1 — SCAN ──────────────────────────────────────────────────────────

def stage1_scan():
    """Scan all emails, extract From headers, filter, deduplicate.
    Returns dict: email_addr -> {name, email, best_body_file, best_body_len}
    """
    if not os.path.exists(INBOX_DIR):
        print(f"Fehler: {INBOX_DIR} existiert nicht.")
        sys.exit(1)

    eml_files = [f for f in os.listdir(INBOX_DIR) if f.endswith('.eml')]
    total = len(eml_files)
    print(f"Scanne {total} E-Mails (Stufe 1 — kein API-Call)...\n")

    contacts = {}  # email_addr -> {name, email, best_body_file, best_body_len}
    skipped_automated = 0
    skipped_invalid = 0
    skipped_own = 0
    skipped_name_mismatch = 0
    progress_interval = max(1, total // 20)  # Show progress ~20 times

    for i, fname in enumerate(eml_files):
        if (i + 1) % progress_interval == 0 or i == total - 1:
            pct = int((i + 1) / total * 100)
            print(f"  Scan: {i+1}/{total} ({pct}%)", end='\r')

        fpath = os.path.join(INBOX_DIR, fname)
        try:
            with open(fpath, 'rb') as f:
                msg = email.message_from_bytes(f.read())
        except Exception:
            skipped_invalid += 1
            continue

        # Parse From header
        from_raw = decode_header(msg.get('From', ''))
        from_name, from_email_addr = email.utils.parseaddr(from_raw)
        from_email_addr = from_email_addr.lower().strip()

        if not from_email_addr or '@' not in from_email_addr:
            skipped_invalid += 1
            continue

        # Filter automated senders
        if is_automated_sender(from_email_addr):
            skipped_automated += 1
            continue

        # Eigene Adressen: NIE als fremder Kontakt — sonst pollutet vCard-Import die My Card.
        if from_email_addr in OWN_EMAILS:
            skipped_own += 1
            continue

        # Name-Sanity: From-Name nur akzeptieren wenn er zur E-Mail-Adresse passt.
        # Sonst Name verwerfen — der Kontakt landet ohne Name in der Vorschau.
        if from_name and not name_matches_email(from_name, from_email_addr):
            skipped_name_mismatch += 1
            from_name = ''

        # Get body length for this email
        body = get_body_text(msg)
        body_len = len(body) if body else 0

        if from_email_addr in contacts:
            # Update name nur wenn neuer Name strict extension UND zur E-Mail passt.
            existing = contacts[from_email_addr]
            existing_name = existing.get('name', '')
            if (from_name
                    and len(from_name) > len(existing_name)
                    and is_strict_name_extension(existing_name, from_name)):
                existing['name'] = from_name
            # Keep reference to email with longest body
            if body_len > existing.get('best_body_len', 0):
                existing['best_body_file'] = fname
                existing['best_body_len'] = body_len
        else:
            contacts[from_email_addr] = {
                'name': from_name or '',
                'email': from_email_addr,
                'best_body_file': fname,
                'best_body_len': body_len,
            }

    print(f"\n\n{'='*60}")
    print(f"  Gesamt gescannt:        {total} E-Mails")
    print(f"  Automatisierte gefiltert: {skipped_automated}")
    print(f"  Eigene Adresse uebersp.: {skipped_own}")
    print(f"  Name-Mismatch verworfen: {skipped_name_mismatch}")
    print(f"  Ungueltig/Fehler:        {skipped_invalid}")
    print(f"  Eindeutige Kontakte:     {len(contacts)}")
    print(f"{'='*60}\n")

    return contacts


def save_preview_excel(contacts):
    """Save preview Excel with just Name + Email."""
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    datestamp = datetime.datetime.now().strftime("%Y-%m-%d")
    xlsx_path = os.path.join(OUTPUT_DIR, f"contacts_preview_{datestamp}.xlsx")

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill

        wb = Workbook()
        ws = wb.active
        ws.title = "Kontakte (Vorschau)"

        header_fill = PatternFill(start_color='1B2A6B', end_color='1B2A6B', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=11)

        for col, h in enumerate(['Name', 'Email'], 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font

        sorted_contacts = sorted(contacts.values(), key=lambda c: c.get('name', '').lower())
        for row_idx, c in enumerate(sorted_contacts, 2):
            ws.cell(row=row_idx, column=1, value=c.get('name', ''))
            ws.cell(row=row_idx, column=2, value=c.get('email', ''))

        # Auto-width
        for col in ws.columns:
            max_len = 0
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

        wb.save(xlsx_path)
        print(f"Vorschau-Excel gespeichert: {xlsx_path}")
        subprocess.Popen(['open', xlsx_path])
        return xlsx_path
    except ImportError:
        print("openpyxl nicht installiert. Installieren: pip3 install openpyxl")
        return None


# ─── STUFE 2 — SIGNATUR-EXTRAKTION ───────────────────────────────────────────

def extract_signatures_batch(api_key, signatures):
    """Send a batch of up to 20 signatures to Claude Haiku, return parsed contacts."""
    from anthropic import Anthropic

    # Build numbered signature list
    sig_parts = []
    for idx, sig in enumerate(signatures, 1):
        sig_parts.append(f"[{idx}]\n{sig['signature_text']}")

    user_message = (
        "Extrahiere aus diesen Signaturen Name, Email, Telefon, Mobil, "
        "Unternehmen, Jobtitel, Website:\n\n"
        + "\n\n".join(sig_parts)
    )

    system_prompt = (
        "Du extrahierst Kontaktdaten aus E-Mail-Signaturen. "
        "Antworte NUR mit einem JSON Array. "
        "Jedes Element hat die Felder: name, email, phone, mobile, company, title, website. "
        "Falls kein Wert gefunden: null. "
        "Die Reihenfolge muss den nummerierten Signaturen entsprechen."
    )

    client = Anthropic(api_key=api_key)
    try:
        response = client.messages.create(
            model="claude-haiku-4-5-20251001",
            max_tokens=2000,
            system=system_prompt,
            messages=[{"role": "user", "content": user_message}],
        )
        text = response.content[0].text.strip()
        # Extract JSON array
        match = re.search(r'\[.*\]', text, re.DOTALL)
        if match:
            results = json.loads(match.group())
            return results
    except json.JSONDecodeError as e:
        print(f"    JSON Parse Fehler: {e}")
    except Exception as e:
        print(f"    API Fehler: {e}")
    return None


def stage2_extract(contacts, continue_mode=False):
    """Stage 2: Extract signature data via Claude Haiku API."""
    api_key = load_api_key()

    # Prepare list of contacts needing processing
    contact_list = sorted(contacts.values(), key=lambda c: c.get('name', '').lower())

    # Load checkpoint if continuing
    processed_emails = set()
    enriched_data = {}
    if continue_mode:
        checkpoint = load_checkpoint()
        if checkpoint and checkpoint.get('stage') == 2:
            processed_emails = set(checkpoint.get('processed_emails', []))
            enriched_data = checkpoint.get('enriched_data', {})
            print(f"Checkpoint geladen: {len(processed_emails)} bereits verarbeitet.\n")
        else:
            print("Kein gueltiger Checkpoint gefunden. Starte von vorn.\n")

    # Filter out already processed
    remaining = [c for c in contact_list if c['email'] not in processed_emails]

    if not remaining:
        print("Alle Kontakte bereits verarbeitet!")
        return enriched_data

    # Prepare signature texts
    sig_items = []
    for c in remaining:
        fname = c.get('best_body_file', '')
        if not fname:
            # No body available, skip but mark as processed
            processed_emails.add(c['email'])
            enriched_data[c['email']] = {
                'name': c.get('name', ''), 'email': c['email'],
                'phone': None, 'mobile': None, 'company': None,
                'title': None, 'website': None,
            }
            continue

        fpath = os.path.join(INBOX_DIR, fname)
        try:
            with open(fpath, 'rb') as f:
                msg = email.message_from_bytes(f.read())
            body = get_body_text(msg)
        except Exception:
            body = ''

        if not body or len(body.strip()) < 30:
            processed_emails.add(c['email'])
            enriched_data[c['email']] = {
                'name': c.get('name', ''), 'email': c['email'],
                'phone': None, 'mobile': None, 'company': None,
                'title': None, 'website': None,
            }
            continue

        # Last 60 lines as signature candidate
        lines = body.strip().split('\n')
        signature_text = '\n'.join(lines[-60:])[:3000]

        sig_items.append({
            'email': c['email'],
            'name': c.get('name', ''),
            'signature_text': signature_text,
        })

    total_batches = (len(sig_items) + BATCH_SIZE - 1) // BATCH_SIZE
    print(f"Stufe 2: {len(sig_items)} Signaturen in {total_batches} Batches\n")

    contacts_with_data = sum(1 for v in enriched_data.values()
                             if any(v.get(k) for k in ('phone', 'mobile', 'company', 'title', 'website')))
    start_time = time.time()

    for batch_idx in range(total_batches):
        batch_start = batch_idx * BATCH_SIZE
        batch_end = min(batch_start + BATCH_SIZE, len(sig_items))
        batch = sig_items[batch_start:batch_end]

        # Progress + ETA
        elapsed = time.time() - start_time
        if batch_idx > 0:
            avg_per_batch = elapsed / batch_idx
            remaining_time = avg_per_batch * (total_batches - batch_idx)
            eta_str = f" | ETA: {format_time(remaining_time)}"
        else:
            eta_str = ""

        print(f"  Batch {batch_idx+1}/{total_batches} | "
              f"Kontakte mit Daten: {contacts_with_data}{eta_str}")

        results = extract_signatures_batch(api_key, batch)

        if results and len(results) == len(batch):
            for item, result in zip(batch, results):
                # Haiku-Name nur akzeptieren wenn er zur E-Mail passt — sonst Header-Name (oder leer)
                haiku_name = result.get('name') or ''
                if haiku_name and not name_matches_email(haiku_name, item['email']):
                    haiku_name = ''
                data = {
                    'name': haiku_name or item['name'],
                    'email': item['email'],
                    'phone': result.get('phone'),
                    'mobile': result.get('mobile'),
                    'company': result.get('company'),
                    'title': result.get('title'),
                    'website': result.get('website'),
                }
                enriched_data[item['email']] = data
                processed_emails.add(item['email'])
                if any(data.get(k) for k in ('phone', 'mobile', 'company', 'title', 'website')):
                    contacts_with_data += 1
        elif results:
            # Partial results — process what we got, mark rest as processed
            for i, item in enumerate(batch):
                if i < len(results):
                    result = results[i]
                    data = {
                        'name': result.get('name') or item['name'],
                        'email': item['email'],
                        'phone': result.get('phone'),
                        'mobile': result.get('mobile'),
                        'company': result.get('company'),
                        'title': result.get('title'),
                        'website': result.get('website'),
                    }
                    enriched_data[item['email']] = data
                    if any(data.get(k) for k in ('phone', 'mobile', 'company', 'title', 'website')):
                        contacts_with_data += 1
                else:
                    enriched_data[item['email']] = {
                        'name': item['name'], 'email': item['email'],
                        'phone': None, 'mobile': None, 'company': None,
                        'title': None, 'website': None,
                    }
                processed_emails.add(item['email'])
        else:
            # API call failed — mark batch items but with empty data
            print(f"    Batch {batch_idx+1} fehlgeschlagen — uebersprungen.")
            for item in batch:
                enriched_data[item['email']] = {
                    'name': item['name'], 'email': item['email'],
                    'phone': None, 'mobile': None, 'company': None,
                    'title': None, 'website': None,
                }
                processed_emails.add(item['email'])

        # Save checkpoint after each batch
        save_checkpoint({
            'stage': 2,
            'processed_emails': list(processed_emails),
            'enriched_data': enriched_data,
            'timestamp': datetime.datetime.now().isoformat(),
        })

    elapsed_total = time.time() - start_time
    print(f"\nStufe 2 abgeschlossen in {format_time(elapsed_total)}.")
    print(f"Kontakte mit erweiterten Daten: {contacts_with_data}/{len(enriched_data)}")

    return enriched_data


# ─── FINAL OUTPUT ─────────────────────────────────────────────────────────────

def save_final_output(enriched_data):
    """Save final Excel + vCard files."""
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    datestamp = datetime.datetime.now().strftime("%Y-%m-%d")

    contact_list = sorted(enriched_data.values(), key=lambda c: (c.get('name') or '').lower())

    # A) Excel
    xlsx_path = os.path.join(OUTPUT_DIR, f"contacts_{datestamp}.xlsx")
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill

        wb = Workbook()
        ws = wb.active
        ws.title = "Kontakte"

        headers = ['Name', 'Email', 'Telefon', 'Mobil', 'Unternehmen', 'Jobtitel', 'Website']
        header_fill = PatternFill(start_color='1B2A6B', end_color='1B2A6B', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=11)

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font

        keys = ['name', 'email', 'phone', 'mobile', 'company', 'title', 'website']
        for row_idx, contact in enumerate(contact_list, 2):
            for col_idx, key in enumerate(keys, 1):
                val = contact.get(key)
                ws.cell(row=row_idx, column=col_idx, value=val if val else '')

        # Auto-width
        for col in ws.columns:
            max_len = 0
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

        wb.save(xlsx_path)
        print(f"\nExcel: {xlsx_path}")
    except ImportError:
        print("\nopenpyxl nicht installiert — pip3 install openpyxl")
        xlsx_path = None

    # B) vCard
    vcf_path = os.path.join(OUTPUT_DIR, f"contacts_{datestamp}.vcf")
    vcf_count = 0
    with open(vcf_path, 'w', encoding='utf-8') as f:
        for contact in contact_list:
            if contact.get('name') or contact.get('email'):
                f.write(contact_to_vcard(contact) + '\n')
                vcf_count += 1
    print(f"vCard: {vcf_path} ({vcf_count} Kontakte)")

    # Open Excel
    if xlsx_path:
        subprocess.Popen(['open', xlsx_path])

    # Ask about Apple Contacts import
    print()
    answer = input("In Apple Contacts importieren? (j/n): ").strip().lower()
    if answer in ('j', 'ja', 'y', 'yes'):
        subprocess.Popen(['open', vcf_path])
        print("vCard geoeffnet — Apple Contacts importiert automatisch.")
    else:
        print(f"Spaeter importieren: open {vcf_path}")

    # Cleanup checkpoint
    if os.path.exists(CHECKPOINT_FILE):
        os.remove(CHECKPOINT_FILE)
        print("Checkpoint-Datei bereinigt.")


# ─── MAIN ─────────────────────────────────────────────────────────────────────

def main():
    continue_mode = '--continue' in sys.argv

    # If --continue, try to resume stage 2 directly
    if continue_mode:
        checkpoint = load_checkpoint()
        if checkpoint and checkpoint.get('stage') == 2:
            print("Setze Stufe 2 fort...\n")
            # Need to re-scan stage 1 to get contact list with body references
            contacts = stage1_scan()
            enriched = stage2_extract(contacts, continue_mode=True)
            save_final_output(enriched)
            return
        else:
            print("Kein Checkpoint gefunden. Starte normal.\n")

    # ─── STUFE 1 ──────────────────────────────────────────────────────────────

    contacts = stage1_scan()

    if not contacts:
        print("Keine Kontakte gefunden.")
        return

    # Save preview Excel
    save_preview_excel(contacts)

    # Cost estimate
    total_batches = (len(contacts) + BATCH_SIZE - 1) // BATCH_SIZE
    est_input_cost = total_batches * EST_TOKENS_PER_BATCH * HAIKU_COST_PER_1K_INPUT / 1000
    est_output_cost = total_batches * EST_OUTPUT_PER_BATCH * HAIKU_COST_PER_1K_OUTPUT / 1000
    est_total_cost = est_input_cost + est_output_cost

    print(f"\nStufe 2 starten fuer Signatur-Extraktion?")
    print(f"Geschaetzte Kosten: ~${est_total_cost:.2f} (Haiku, {total_batches} Batches)")
    answer = input("(j/n): ").strip().lower()

    if answer not in ('j', 'ja', 'y', 'yes'):
        print("\nAbgebrochen. Vorschau-Excel wurde gespeichert.")
        print(f"Spaeter fortsetzen: python3 {__file__} --continue")
        # Save checkpoint so --continue knows where we are
        save_checkpoint({
            'stage': 2,
            'processed_emails': [],
            'enriched_data': {},
            'timestamp': datetime.datetime.now().isoformat(),
        })
        return

    # ─── STUFE 2 ──────────────────────────────────────────────────────────────

    enriched = stage2_extract(contacts)

    # ─── FINAL OUTPUT ─────────────────────────────────────────────────────────

    save_final_output(enriched)


if __name__ == '__main__':
    main()
